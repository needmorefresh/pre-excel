import json, couchdb
from openpyxl import load_workbook
from cassandra.cluster import Cluster
from cassandra.auth import PlainTextAuthProvider

file = load_workbook('./excel/pre.xlsx', data_only=True)
s1 = file['Sheet1']

clients = {}

## 0 사업자번호 / 1 사업자명 / 2 KPN건수 / 3 KPN객단가 / 4 KOVAN(포코)건수 / 5 KOVAN(포코)객단가
## 6 KIS(포코)건수 / 7 KIS(포코)객단가 / 8 JTNet건수 / 9 JTNet객단가 / 10 KOVAN(e포)객단가 / 11 KOVAN(e포)객단가
## 12 KIS(e포)건수 / 13 KIS(e포)객단가 / 14 NICE건수 / 15 NICE객단가 / 16 KCP건수 / 17 KCP객단가
## 18 대표자명 / 19 전화번호 / 20 주소
for row in s1.iter_rows():
    marketNumber = row[0].value
    marketName = row[1].value
    ownerName = row[18].value
    location = row[20].value
    tel = row[19].value

    if row[0].coordinate == 'A1':
        continue
    
    if marketNumber not in clients:
        clients[marketNumber] = {}
        clients[marketNumber]['tos'] = {}
        clients[marketNumber]['aov'] = {}
        clients[marketNumber]['marketName'] = marketName
        clients[marketNumber]['ownerName'] = None
        clients[marketNumber]['tel'] = None
        clients[marketNumber]['loc'] = None

    if clients[marketNumber]['ownerName'] == None and ownerName != None:
        clients[marketNumber]['ownerName'] = ownerName
    if clients[marketNumber]['tel'] == None and tel != None:
        clients[marketNumber]['tel'] = '0' + str(tel)
    if clients[marketNumber]['loc'] == None and location != None:
        clients[marketNumber]['loc'] = ' '.join(location.split())

    for i in (2, 4, 6, 8, 10, 12, 14, 16):
        van = {2: 'kpn1', 4: 'kov1', 6: 'kis1', 8: 'jtn1', 10: 'kov2', 12: 'kis2', 14: 'nic1', 16: 'kcp1'}
        if row[i].value != None:
            clients[marketNumber]['tos'][van[i]] = 0 if row[i].value == None else row[i].value
            clients[marketNumber]['aov'][van[i]] = int(''.join(row[i+1].value.split(','))) if row[i+1].value != None else 0

print(f'{next(iter(clients))} : {clients[next(iter(clients))]}')

data = []

for key, value in clients.items():
    data.append(dict(storeId=str(key), tos=value['tos'], aov=value['aov'],
                     storeName=value['marketName'], tel=value['tel'], location=value['loc']))

couch = couchdb.Server("http://127.0.0.1:5984/")
db = couch["withpos_stores"]

for row in data:
    db.save(row)

## Let's Make JSON!!
with open('./rawdata/data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent='\t', ensure_ascii=False)

'''
cluster = Cluster(
    contact_points=[
        "localhost",
    ],
    auth_provider=PlainTextAuthProvider(username='scylla', password='your-awesome-password')
)
session = cluster.connect()
'''